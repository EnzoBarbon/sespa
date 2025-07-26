import camelot  # pip install "camelot‑py[cv]" ghostscript
import pandas as pd
import datetime
import sys
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def extract_situaciones(pdf_path, pages="2-5"):
    """
    Devuelve un DataFrame con cada fila de los cuadros 'SITUACIÓN/ES'
    de un informe de vida laboral.

    Parámetros
    ----------
    pdf_path : str
        Ruta al PDF.
    pages : str
        Rango de páginas (1‑indexado).  Ej.: "2-5", "2,3,4,5".
    """
    tablas = camelot.read_pdf(pdf_path, pages=pages, flavor="stream")
    registros, corriente = [], None

    for tabla in tablas:
        df = tabla.df

        for _, fila_raw in df.iterrows():
            # Normalizamos: quitamos saltos de línea y espacios extra
            fila = [" ".join(str(c).split()) for c in fila_raw.tolist()]

            # Filtramos cabeceras y separadores
            if (fila[0].upper().startswith(("RÉGIMEN", "SITUACIÓN"))
                or fila[0] == "" and not fila[1].isdigit()):
                continue

            # Filas de continuación (la celda 0 viene vacía)
            if fila[0] == "":
                if corriente and fila[2]:
                    corriente["Empresa"] += " " + fila[2]
                continue

            # Rellenamos celdas que falten para no romper el índice
            while len(fila) < 10:
                fila.append("")

            corriente = {
                "Regimen":           fila[0],
                "Codigo_Empresa":    fila[1],
                "Empresa":           fila[2],
                "Fecha_Alta":        fila[3],
                "Fecha_Efecto_Alta": fila[4],
                "Fecha_Baja":        fila[5],
                "C.T.":              fila[6],
                "CTP_%":             fila[7],
                "G.C.":              fila[8],
                "Dias":              fila[9],
            }
            registros.append(corriente)

    return pd.DataFrame(registros)


# === Uso rápido =============================================================
parser = argparse.ArgumentParser(description="Extract situaciones from PDF or images.")
parser.add_argument('--filter-2008', action='store_true', help='Filter rows where Fecha_Alta is before 2008')
parser.add_argument('--use-ocr', action='store_true', help='Use OCR on images instead of PDF processing')
args = parser.parse_args()

if args.use_ocr:
    # Use OCR processing on images
    from ocr_processor import process_all_images, convert_ocr_to_extract_format
    
    print("Using OCR processing on images...")
    ocr_records = process_all_images("data/imagenes")
    output_data = convert_ocr_to_extract_format(ocr_records)
    
    # Create a dummy DataFrame for compatibility with existing code
    df = pd.DataFrame()
    
else:
    # Use PDF processing (original method)
    print("Using PDF processing...")
    pdf_path = "data/data.pdf"         # hardcoded path to the PDF in the data directory
    df = extract_situaciones(pdf_path, pages="2-5")

    # Convert Fecha_Alta to datetime for possible filtering
    # Handle possible empty or malformed dates gracefully
    def parse_date(date_str):
        try:
            return datetime.datetime.strptime(date_str, "%d.%m.%Y")
        except Exception:
            return None

    df["Fecha_Alta_dt"] = df["Fecha_Alta"].apply(parse_date)

    if args.filter_2008:
        df_filtered = df[df["Fecha_Alta_dt"].notnull() & (df["Fecha_Alta_dt"] < datetime.datetime(2008, 1, 1))]
    else:
        df_filtered = df.copy()

    def format_date(date_str):
        try:
            d = datetime.datetime.strptime(date_str, "%d.%m.%Y")
            return d.strftime("%d/%m/%Y")
        except Exception:
            return ""

    # Extract only relevant records directly
    output_data = []
    for _, row in df_filtered.iterrows():
        empresa = row["Empresa"]
        
        # Check if it's a vacation record
        if empresa.startswith("VACACIONES RETRIBUIDAS Y NO"):
            output_data.append({
                "isVacaciones": True,
                "fechaAlta": format_date(row["Fecha_Alta"]),
                "fechaBaja": format_date(row["Fecha_Baja"])
            })
        # Check if it's a health service contract
        elif empresa.startswith("SERVICIO DE SALUD DEL PRINCIPADO"):
            output_data.append({
                "isVacaciones": False,
                "fechaAlta": format_date(row["Fecha_Alta"]),
                "fechaBaja": format_date(row["Fecha_Baja"])
            })

def calculate_non_overlapping_vacation_days(data):
    """
    Calculate vacation days that don't overlap with contract periods.
    Returns both the total days and the specific non-overlapping periods.
    """
    vacations = [item for item in data if item["isVacaciones"]]
    contracts = [item for item in data if not item["isVacaciones"]]
    
    total_non_overlapping_days = 0
    non_overlapping_periods = []
    
    for vacation in vacations:
        if not vacation["fechaAlta"] or not vacation["fechaBaja"]:
            continue
            
        try:
            vac_start = datetime.datetime.strptime(vacation["fechaAlta"], "%d/%m/%Y")
            vac_end = datetime.datetime.strptime(vacation["fechaBaja"], "%d/%m/%Y")
        except ValueError:
            continue
            
        # Find overlapping contracts
        overlapping_periods = []
        
        for contract in contracts:
            if not contract["fechaAlta"]:
                continue
                
            try:
                contract_start = datetime.datetime.strptime(contract["fechaAlta"], "%d/%m/%Y")
                # If contract has no end date, assume it's still active (use today)
                if contract["fechaBaja"]:
                    contract_end = datetime.datetime.strptime(contract["fechaBaja"], "%d/%m/%Y")
                else:
                    contract_end = datetime.datetime.now()
                    
                # Check if vacation overlaps with contract
                if vac_start <= contract_end and vac_end >= contract_start:
                    # Calculate overlapping period
                    overlap_start = max(vac_start, contract_start)
                    overlap_end = min(vac_end, contract_end)
                    overlapping_periods.append((overlap_start, overlap_end))
                    
            except ValueError:
                continue
        
        # Calculate non-overlapping periods for this vacation
        if not overlapping_periods:
            # No overlap, count all vacation days
            vacation_days = (vac_end - vac_start).days + 1
            total_non_overlapping_days += vacation_days
            non_overlapping_periods.append({
                "start": vac_start.strftime("%d/%m/%Y"),
                "end": vac_end.strftime("%d/%m/%Y"),
                "days": vacation_days
            })
        else:
            # Merge overlapping periods and calculate non-overlapping segments
            overlapping_periods.sort()
            merged_overlaps = []
            
            for overlap in overlapping_periods:
                if not merged_overlaps or merged_overlaps[-1][1] < overlap[0]:
                    merged_overlaps.append(overlap)
                else:
                    merged_overlaps[-1] = (merged_overlaps[-1][0], max(merged_overlaps[-1][1], overlap[1]))
            
            # Find non-overlapping segments within this vacation
            current_pos = vac_start
            
            for overlap_start, overlap_end in merged_overlaps:
                # Add period before this overlap (if any)
                if current_pos < overlap_start:
                    segment_end = overlap_start - datetime.timedelta(days=1)
                    segment_days = (segment_end - current_pos).days + 1
                    if segment_days > 0:
                        total_non_overlapping_days += segment_days
                        non_overlapping_periods.append({
                            "start": current_pos.strftime("%d/%m/%Y"),
                            "end": segment_end.strftime("%d/%m/%Y"),
                            "days": segment_days
                        })
                
                # Move current position past this overlap
                current_pos = overlap_end + datetime.timedelta(days=1)
            
            # Add remaining period after all overlaps (if any)
            if current_pos <= vac_end:
                segment_days = (vac_end - current_pos).days + 1
                if segment_days > 0:
                    total_non_overlapping_days += segment_days
                    non_overlapping_periods.append({
                        "start": current_pos.strftime("%d/%m/%Y"),
                        "end": vac_end.strftime("%d/%m/%Y"),
                        "days": segment_days
                    })
    
    return total_non_overlapping_days, non_overlapping_periods

def create_excel_report(total_days, vacation_periods, filename="vacation_report.xlsx"):
    """
    Create an Excel report with vacation summary and detailed periods.
    """
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    
    # Header styling
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    
    # Title
    ws_summary['A1'] = "RESUMEN DE VACACIONES NO SOLAPADAS"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:C1')
    ws_summary['A1'].alignment = header_alignment
    
    # Total days
    ws_summary['A3'] = "Total de días de vacaciones no solapados:"
    ws_summary['A3'].font = Font(bold=True)
    ws_summary['B3'] = total_days
    ws_summary['B3'].font = Font(size=14, bold=True)
    
    # Number of periods
    ws_summary['A4'] = "Número de períodos:"
    ws_summary['A4'].font = Font(bold=True)
    ws_summary['B4'] = len(vacation_periods)
    ws_summary['B4'].font = Font(size=14)
    
    # Detailed periods sheet
    ws_periods = wb.create_sheet("Períodos Detallados")
    
    # Headers for periods
    headers = ["Fecha Inicio", "Fecha Fin", "Días"]
    for col, header in enumerate(headers, 1):
        cell = ws_periods.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows - one row per vacation period
    for row, period in enumerate(vacation_periods, 2):
        ws_periods.cell(row=row, column=1, value=period["start"])
        ws_periods.cell(row=row, column=2, value=period["end"])
        ws_periods.cell(row=row, column=3, value=period["days"])
    
    # Add total row
    if vacation_periods:
        total_row = len(vacation_periods) + 3
        ws_periods.cell(row=total_row, column=2, value="TOTAL:").font = Font(bold=True)
        ws_periods.cell(row=total_row, column=3, value=total_days).font = Font(bold=True)
    
    # Adjust column widths
    from openpyxl.utils import get_column_letter
    for ws in [ws_summary, ws_periods]:
        for col_num in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max(min(max_length + 2, 50), 10)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    print(f"Excel report saved as: {filename}")
    return filename

import json
with open("output/py_output.json", "w", encoding="utf-8") as f:
    json.dump(output_data, f, ensure_ascii=False, indent=2)

# Calculate non-overlapping vacation days
non_overlapping_vacation_days, vacation_periods = calculate_non_overlapping_vacation_days(output_data)
print(f"Total non-overlapping vacation days: {non_overlapping_vacation_days}")
print(f"Non-overlapping vacation periods: {len(vacation_periods)} periods")

# Save results with additional info
results = {
    "data": output_data,
    "total_non_overlapping_vacation_days": non_overlapping_vacation_days,
    "non_overlapping_vacation_periods": vacation_periods
}

with open("output/py_output_with_calculation.json", "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=2)

# Create Excel report
excel_filename = "output/vacation_report.xlsx"
create_excel_report(non_overlapping_vacation_days, vacation_periods, excel_filename)

# Ejemplos de salida
if not args.use_ocr and not df.empty:
    print(df.head())        # primeras filas
    df.to_csv("situaciones.csv", index=False) # exportar a CSV
else:
    print(f"OCR processing completed. {len(output_data)} records processed.")
